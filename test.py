import traceback

from openpyxl import load_workbook
from sheetfu import SpreadsheetApp, Table
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
from selenium.webdriver import EdgeOptions as Options
from selenium.webdriver import EdgeService as Service
from subprocess import CREATE_NO_WINDOW
import time
import datetime
from datetime import datetime
import queue
import tkinter as tk
from threading import Thread
from queue import Queue

status_queue = Queue()


def log(data):
    wb = load_workbook("IndentLog.xlsx")
    ws = wb.active
    date = datetime.now().strftime("%d-%m-%Y %H:%M.%S")
    ws.cell(ws.max_row + 1, 1).value = data
    ws.cell(ws.max_row + 1, 2).value = date
    wb.save("IndentLog.xlsx")


def data_exists(table, data):
    status_queue.put(f"Checking Indent: {data['Indent']} for availability")
    for item in table:
        if data["Indent"] == item.get_field_value("Indent"):
            print("Data Exists")
            status_queue.put(f"Indent: {data['Indent']} already exists")
            return True
    return False


def add_data(data):
    print(data)
    sa = SpreadsheetApp('encoded-net-397411-ac78804e5820.json')
    spreadsheet = sa.open_by_id('1dIC0YJrQn2YdiJawl-etvttB83PK-ZQfWcM0zHkNKe8')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    print(data_range)
    table = Table(data_range)
    if data_exists(table, data):
        return
    table.add_one(data)
    table.commit()


def get_remaining_data(script):
    service = Service()
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 20)
    action = webdriver.ActionChains(driver)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    driver.execute_script(script)
    qty = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div[3]/div/table/tfoot/tr/td[2]'))).get_attribute("innerText")
    table = driver.find_element(By.XPATH, '/html/body/form/div[4]/div[2]/table/tbody/tr/td/table/tbody')
    last_approver = table.find_elements(By.CSS_SELECTOR, "td[style='width:350px;     vertical-align: middle;']")[-1].get_attribute("innerText").split("(")[0]
    print("last approver: " + last_approver)
    return qty, last_approver


def run_and_approve():
    service = Service()
    options = Options()
    # options.add_argument("--headless")
    # options.add_argument("--no-sandbox")
    # options.add_argument("--window-size=1280,720")
    # options.add_argument("--disable-gpu")
    # service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 20)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    table = driver.find_element(By.ID, "tablea").find_element(By.TAG_NAME, "tbody")
    rows = table.find_elements(By.TAG_NAME, "tr")

    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")
        data = {
            "Indent": tds[0].get_attribute("innerText"),
        }
        # script = tds[10].find_element(By.TAG_NAME, "input").get_attribute("onclick")
        # print(script)

        if data_exists(table, data):
            continue

            data["Approval_Status"] = "Approved"
            data["Bot_Updated_Status"] = "Updated"
            data["Bot_Updated_Dt"] = datetime.datetime.now().strftime("%d-%m-%Y %H:%M")
            data["UpdatedBy_Sys"] = "Bot Skipped"

        add_data(data)

run_and_approve()
