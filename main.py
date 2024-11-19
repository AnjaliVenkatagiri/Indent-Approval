import traceback

from openpyxl import load_workbook
from sheetfu import SpreadsheetApp, Table
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options
from selenium.webdriver import EdgeOptions as Options
from selenium.webdriver import EdgeService as Service
from subprocess import CREATE_NO_WINDOW
import time
import datetime
from datetime import datetime
import queue
import tkinter as tk
from threading import Thread
from queue import Queue

status_queue = Queue()


def log(data):
    wb = load_workbook("IndentLog.xlsx")
    ws = wb.active
    date = datetime.now().strftime("%d-%m-%Y %H:%M.%S")
    ws.cell(ws.max_row + 1, 1).value = data
    ws.cell(ws.max_row + 1, 2).value = date
    wb.save("IndentLog.xlsx")


def data_exists(table, data):
    status_queue.put(f"Checking Indent: {data['Indent']} for availability")
    for item in table:
        if data["Indent"] == item.get_field_value("Indent"):
            print("Data Exists")
            status_queue.put(f"Indent: {data['Indent']} already exists")
            return True
    return False


def add_data(data):
    print(data)
    sa = SpreadsheetApp('encoded-net-397411-ac78804e5820.json')
    spreadsheet = sa.open_by_id('1dIC0YJrQn2YdiJawl-etvttB83PK-ZQfWcM0zHkNKe8')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    print(data_range)
    table = Table(data_range)
    if data_exists(table, data):
        return
    table.add_one(data)
    table.commit()


def get_remaining_data(script):
    service = Service()
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 20)
    action = webdriver.ActionChains(driver)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    driver.execute_script(script)
    qty = wait.until(ec.presence_of_element_located((By.XPATH, '/html/body/form/div[3]/div/table/tfoot/tr/td[2]'))).get_attribute("innerText")
    table = driver.find_element(By.XPATH, '/html/body/form/div[4]/div[2]/table/tbody/tr/td/table/tbody')
    last_approver = table.find_elements(By.CSS_SELECTOR, "td[style='width:350px;     vertical-align: middle;']")[-1].get_attribute("innerText").split("(")[0]
    print("last approver: " + last_approver)
    return qty, last_approver


def approve(indent):
    service = Service()
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 20)
    action = webdriver.ActionChains(driver)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    script = f"javascript:window.location.href='IndentMaster.jsp?reqno={indent}&revno=0&myusrid=901519'"
    driver.execute_script(script)
    try:
        remark = wait.until(ec.presence_of_element_located((By.TAG_NAME, 'textarea')))
        driver.execute_script("arguments[0].value = 'APPROVED .';", remark)
        forward_to = driver.find_element(By.NAME, "forward_to")
        driver.execute_script("arguments[0].value = '299774';", forward_to)
        button = driver.find_element(By.NAME, "btn")
        driver.execute_script("arguments[0].click();", button)
        return True
    except:
        return False


def scrapper():
    status_queue.put("Scrapper bot started")
    service = Service()
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    wait = WebDriverWait(driver, 20)
    action = webdriver.ActionChains(driver)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    table = driver.find_element(By.ID, "tablea").find_element(By.TAG_NAME, "tbody")
    rows = table.find_elements(By.TAG_NAME, "tr")
    for i in rows:
        tds = i.find_elements(By.TAG_NAME, "td")
        data = {
            "Indent": tds[0].get_attribute("innerText"),
            "Date": str(datetime.strptime(tds[1].get_attribute("innerText"), "%d/%m/%Y").strftime("%m/%d/%Y")),
            "PerformaInvoiceNo": tds[2].get_attribute("innerText"),
            "PIDate": str(datetime.strptime(tds[3].get_attribute("innerText"), "%d/%m/%Y").strftime("%m/%d/%Y")),
            "Supplier": tds[4].get_attribute("innerText"),
            "Curr": tds[5].get_attribute("innerText"),
            "Amount": tds[6].get_attribute("innerText"),
            "Description": tds[7].get_attribute("innerText"),
            "RequestedBy": tds[8].get_attribute("innerText"),
            "ForwardTo": tds[9].get_attribute("innerText"),
        }
        script = tds[10].find_element(By.TAG_NAME, "input").get_attribute("onclick")
        print(data)
        data["Qty"], data["LastApprover"] = get_remaining_data(script)
        add_data(data)


def search_and_approve():
    status_queue.put("Approver bot started")
    sa = SpreadsheetApp('encoded-net-397411-ac78804e5820.json')
    spreadsheet = sa.open_by_id('1dIC0YJrQn2YdiJawl-etvttB83PK-ZQfWcM0zHkNKe8')
    sheet = spreadsheet.get_sheet_by_name('Sheet1')
    data_range = sheet.get_data_range()
    table = Table(data_range)
    for item in table:
        status_queue.put(f"Checking Approval for Indent: {item.get_field_value('Indent')}")
        if "Approve" in item.get_field_value("Approval_Status") and "Updated" not in item.get_field_value("Bot_Updated_Status"):
            try:
                if approve(item.get_field_value("Indent")):
                    item.set_field_value("Bot_Updated_Status", "Updated")
                    item.set_field_value("Bot_Updated_Dt", datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))
                    status_queue.put(f"Approved for Indent: {item.get_field_value('Indent')}")
                    table.commit()
                else:
                    item.set_field_value("Bot_Updated_Status", "Updated")
                    item.set_field_value("Bot_Updated_Dt", datetime.datetime.now().strftime("%d-%m-%Y %H:%M"))
                    item.set_field_value("UpdatedBy_Sys", "Bot Skipped")
                    status_queue.put(f"Bot skipped approval for Indent: {item.get_field_value('Indent')}")
                    table.commit()
            except:
                traceback.print_exc()
                status_queue.put(f"Error Approving {item.get_field_value('Indent')}")


def run_scraper():
    try:
        scrapper()
    except Exception as e:
        print(f"Some Issue Scrapping data {repr(e)}")
        traceback.print_exc()
        status_queue.put(f"Some Issue Scrapping data {repr(e)} rerunning after 30min")
    finally:
        time.sleep(900)
        run_scraper()


def run_approver():
    try:
        search_and_approve()
    except Exception as e:
        print(f"Some Issue Approving data {repr(e)}")
        print(traceback.format_exc())
        status_queue.put(f"Some Issue Approving data {repr(e)} rerunning after 30min")
    finally:
        time.sleep(900)
        run_approver()


def indent_exists_on_webpage(indent_number):
    service = Service()
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1280,720")
    options.add_argument("--disable-gpu")
    # options.add_experimental_option("detach", True)
    service.creation_flags = CREATE_NO_WINDOW
    driver = webdriver.Edge(service=service, options=options)
    driver.get("http://172.17.3.57:8080/IntraNet/Imports/IndentAprovalQry.jsp?myusrid=901519")
    table = driver.find_element(By.ID, "tablea").find_element(By.TAG_NAME, "tbody")
    rows = table.find_elements(By.TAG_NAME, "tr")
    for row in rows:
        tds = row.find_elements(By.TAG_NAME, "td")
        data = tds[0].get_attribute("innerText")
        if indent_number in data:
            return True
    return False


def run_check():
    try:
        sa = SpreadsheetApp('encoded-net-397411-ac78804e5820.json')
        spreadsheet = sa.open_by_id('1dIC0YJrQn2YdiJawl-etvttB83PK-ZQfWcM0zHkNKe8')
        sheet = spreadsheet.get_sheet_by_name('Sheet1')
        data_range = sheet.get_data_range()
        table = Table(data_range)

        for item in table:
            indent_number = item.get_field_value('Indent')
            status_queue.put(f"Checking Indent for Availability: {indent_number}")
            if "Approve" in item.get_field_value("Approval_Status"):
                continue
            if not indent_exists_on_webpage(indent_number):
                item.set_field_value("Approval_Status", "Approved")
                item.set_field_value("Bot_Updated_Status", "Updated")
                item.set_field_value("Bot_Updated_Dt", datetime.now().strftime("%d-%m-%Y %H:%M"))
                item.set_field_value("UpdatedBy_Sys", "Bot skipped")
                status_queue.put(f"Bot skipped approval for Indent: {indent_number}")
                table.commit()
    except Exception as e:
        print(f"Some Issue Checking Indent data {repr(e)}")
        print(traceback.format_exc())
        # status_queue.put(f"Some Issue Approving data {repr(e)} rerunning after 30min")
    finally:
        time.sleep(900)
        run_check()


def update_label():
    try:
        status = status_queue.get(timeout=3)
        label.config(text=status)
        log(status)
    except queue.Empty:
        label.config(text="Waiting")
        log("Waiting")
    finally:
        root.after(1000, update_label)


# approve("2002374437")
thread1 = Thread(target=run_scraper, daemon=True)
thread2 = Thread(target=run_approver, daemon=True)
thread3 = Thread(target=run_check, daemon=True)
thread1.start()
thread2.start()
thread3.start()
root = tk.Tk()
root.geometry("700x200")
root.title("Indent Approval")
label = tk.Label(root, text="Application started")
label.pack(pady=70)
root.after(1000, update_label)
root.mainloop()

# scrapper()

